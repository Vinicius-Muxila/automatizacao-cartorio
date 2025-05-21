import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Lista com os nomes dos meses
meses = ["jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]

# Dicionário para armazenar os dados
dados_faturamento = {}

# Ler cada planilha e consolidar os dados
for mes in meses:
    nome_arquivo = f"planilha_{mes}.xlsx"  # Nome do arquivo
    df = pd.read_excel(nome_arquivo)  # Ler a planilha
    
    # Garantindo que as colunas estejam sempre em maiúsculas
    df.columns = df.columns.str.upper()
    
    # Converter a coluna VALOR para float (caso esteja como string)
    df["VALOR"] = pd.to_numeric(df["VALOR"], errors='coerce').fillna(0)
    
    # Somar faturamento por empresa no mês
    df = df.groupby("CLIENTE")["VALOR"].sum().reset_index()
    
    # Adicionar os dados ao dicionário
    for index, row in df.iterrows():
        cliente = row["CLIENTE"]
        valor = row["VALOR"]
        
        if cliente not in dados_faturamento:
            dados_faturamento[cliente] = {mes: 0 for mes in meses}  # Inicializa todos os meses com 0
        
        dados_faturamento[cliente][mes] = valor

# Criar DataFrame final
clientes = list(dados_faturamento.keys())
dados_processados = []

for cliente, valores in dados_faturamento.items():
    linha = {"CLIENTE": cliente}
    linha.update(valores)
    
    # Calcular faturamento total
    faturamento_total = sum(map(float, valores.values()))
    
    # Calcular quantos meses a empresa faturou
    meses_faturados = sum(1 for v in valores.values() if v > 0)
    linha["FATURAMENTO_TOTAL"] = faturamento_total
    linha["MESES_FATURADOS"] = f"{meses_faturados}/12"
    linha["FATURAMENTO_MEDIO"] = faturamento_total / 12  # Média baseada em 12 meses
    
    dados_processados.append(linha)

df_final = pd.DataFrame(dados_processados)

# Estatísticas gerais
total_empresas = len(df_final)
empresas_acima_media = df_final[df_final["FATURAMENTO_MEDIO"] >= 300].shape[0]
empresas_abaixo_media = total_empresas - empresas_acima_media
percentual_acima_media = (empresas_acima_media / total_empresas) * 100 if total_empresas > 0 else 0
faturamento_total_geral = df_final["FATURAMENTO_TOTAL"].sum()

# Criar DataFrame com estatísticas
df_estatisticas = pd.DataFrame({
    "MÉTRICA": ["Total de Empresas", "Empresas Acima da Média", "Empresas Abaixo da Média", "Percentual Acima da Média", "Faturamento Total"],
    "VALOR": [total_empresas, empresas_acima_media, empresas_abaixo_media, f"{percentual_acima_media:.2f}%", f"R$ {faturamento_total_geral:,.2f}"]
})

# Criar um writer para salvar as informações em diferentes abas
arquivo_saida = "faturamento_anual.xlsx"
with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
    df_estatisticas.to_excel(writer, sheet_name="Estatísticas", index=False)
    df_final.to_excel(writer, sheet_name="Faturamento Consolidado", index=False)

# Aplicar formatação na planilha
wb = openpyxl.load_workbook(arquivo_saida)
ws = wb["Faturamento Consolidado"]

# Ajustar cabeçalho
header_fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_style

# Ajustar largura das colunas
def ajustar_largura(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

ajustar_largura(ws)

# Formatar os valores monetários
for row in ws.iter_rows(min_row=2, min_col=2, max_col=14):  # Colunas de valores
    for cell in row:
        cell.number_format = "R$ #,##0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = border_style

# Ajustar aba de estatísticas
ws_stats = wb["Estatísticas"]
ajustar_largura(ws_stats)

for cell in ws_stats[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_style

for row in ws_stats.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

# Salvar a planilha formatada
wb.save(arquivo_saida)

print("Planilha consolidada gerada e formatada com sucesso!")

